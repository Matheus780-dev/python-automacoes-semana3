"""
Projeto Relatório em Excel
Objetivo

Agrupar vendas por Cliente ou Produto

Calcular totais (soma do valor gasto)

Exportar para Excel com data/hora no nome
"""
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# dados fictícios de vendas
dados = {
    "Data": [
        datetime(2025, 8, 1),
        datetime(2025, 8, 2),
        datetime(2025, 8, 2),
        datetime(2025, 8, 3),
        datetime(2025, 8, 3),
        datetime(2025, 8, 4),
    ],
    "Cliente": ["Maria", "João", "Ana", "Pedro", "Maria", "João"],
    "Produto": ["Notebook", "Mouse", "Teclado", "Monitor", "Mouse", "Headset"],
    "Quantidade": [1, 2, 1, 1, 1, 2],
    "Preço": [3500, 80, 120, 900, 80, 150],
}

df = pd.DataFrame(dados)

#  coluna Total
df["Total"] = df["Quantidade"] * df["Preço"]

print(df)

# Agrupar dados
relatorio_cliente = df.groupby("Cliente")["Total"].sum().reset_index()
print(relatorio_cliente)  # total gasto por cliente

relatorio_produto = df.groupby("Produto")["Total"].sum().reset_index()
print(relatorio_produto)  # total vendido por produto

# exportar para excel com data/hora no nome

nome_arquivo = f"relatorio_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

# Exportando para Excel com 3 abas
with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:
    relatorio_cliente.to_excel(writer, sheet_name="Por Cliente", index=False)
    relatorio_produto.to_excel(writer, sheet_name="Por Produto", index=False)
    df.to_excel(writer, sheet_name="Detalhado", index=False)

#  Abrir com openpyxl para formatar colunas
wb = load_workbook(nome_arquivo)

for sheet in ["Por Cliente", "Por Produto", "Detalhado"]:
    ws = wb[sheet]

    # Procurar colunas com números (Preço, Total)
    for col in ws.iter_cols(min_row=2):  # começa na linha 2 (ignora cabeçalho)
        for cell in col:
            if isinstance(cell.value, (int, float)):  # só números
                cell.number_format = u'R$ #,##0.00'  # formato moeda BRL

# Salvar arquivo formatado
wb.save(nome_arquivo)

print(f"relatorio salvo como {nome_arquivo}")
